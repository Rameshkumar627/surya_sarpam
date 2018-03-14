# -*- coding: utf-8 -*-

from odoo import fields, api, exceptions, _
from datetime import datetime
from .. import surya
import json
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from StringIO import StringIO
import base64
import cStringIO


REPORT_TYPE = [('html', 'HTML'), ('pdf', 'PDF'), ('excel', 'Excel')]


class Report(surya.Sarpam):
    _name = "report.report"

    name = fields.Char(string="Name")
    others = fields.Char(string="Others")

    report_type = fields.Selection(selection=REPORT_TYPE, string="Report Type")

    column_detail = fields.One2many(comodel_name="report.column",
                                    inverse_name="report_id",
                                    string="Data Field")
    report_design = fields.One2many(comodel_name="report.design",
                                    inverse_name="report_id",
                                    string="Report Design")

    model = fields.Many2one(comodel_name="ir.model", string="Model")
    sorting = fields.Many2one(comodel_name="ir.model.fields", string="Sorting")
    search_detail = fields.One2many(comodel_name="report.search", string="Search Detail", inverse_name="search_id")
    filename = fields.Char(string="File name")
    file_output = fields.Binary(string="File Output")

    header_row = fields.Integer(string="Header Row")
    body_row = fields.Integer(string="Body Row")
    footer_row = fields.Integer(string="Footer Row")

    @api.multi
    def data_query(self):
        detail = []

        for rec in self.search_detail:
            detail.append((rec.data_field, rec.operator, rec.value))

        records = self.env[self.model.model].search(detail)
        recs = self.column_detail.sorted(key=lambda r: r.column_seq)

        row = []
        for record in records:
            column = []

            for rec in recs:
                print eval("record.{0}".format(rec.data_field))
                column.append(eval("record.{0}".format(rec.data_field)))
            row.append(column)

        return row

    def above_details(self, ws):
        recs = self.report_design.sorted(key=lambda r: r.order_seq)
        for rec in recs:
            ws.merge_cells(rec.row_col)
            ws[rec.row_col[:rec.row_col.find(":")]] = rec.name
        return ws

    def header_details(self):
        row = []

        recs = self.column_detail.sorted(key=lambda r: r.column_seq)
        for rec in recs:
            row.append(rec.name)

        return row

    def footer_details(self, records):
        row = []

        recs = self.column_detail.sorted(key=lambda r: r.column_seq)
        for rec in recs:
            if rec.sum_footer:
                row.append(0)
            else:
                row.append(None)

        for record in records:
            for rec in recs:
                if rec.sum_footer:
                    row[rec.column_seq] = row[rec.column_seq] + record[rec.column_seq]

        return row

    def below_details(self):
        pass

    def styling(self):
        pass

    @api.multi
    def trigger_field_data(self):
        wb = Workbook()
        ws = wb.active

        ws = self.above_details(ws)

        header = self.header_details()
        ws.append(header)

        bodies = self.data_query()
        for body in bodies:
            ws.append(body)

        footer = self.footer_details(bodies)
        ws.append(footer)

        self.below_details()



        output = cStringIO.StringIO()
        wb.save(output)
        out = base64.encodestring(output.getvalue())

        self.file_output = out

    def row_creation_above(self, sheet1):
        recs = self.env["report.design"].search([("report_id", "=", self.id),
                                                 ("row_type", "=", "above")])

        recs.sorted(key=lambda r: r.order_seq)

        for rec in recs:
            style = xlwt.easyxf(rec.style_id.style)
            sheet1.write_merge(rec.row_1, rec.row_2, rec.col_1, rec.col_2, rec.data, style)

        return sheet1

    def header_creation(self, sheet1):
        rec = self.env["report.design"].search([("report_id", "=", self.id),
                                                ("row_type", "=", "header")])

        row = rec.start_row

        columns = self.env["report.column"].search([("column_seq", ">", 0)])
        style = xlwt.easyxf(rec.style_id.style)

        for column in columns:
            sheet1.write(row, column.column_seq, column.name, style)

        return sheet1, row


class ReportSearch(surya.Sarpam):
    _name = "report.search"

    data_field = fields.Char(string="Data Field")
    operator = fields.Selection(selection=[("=", "=")], string="Operator")
    value = fields.Char(string="Value")
    search_id = fields.Many2one(comodel_name="report.report", string="Search")


class ReportColumn(surya.Sarpam):
    _name = "report.column"

    name = fields.Char(string="Name")
    data_field = fields.Char(string="Data Field")
    column_seq = fields.Integer(string="Column")
    sum_footer = fields.Boolean(string="Sum")

    header_style_excel = fields.Many2one(comodel_name="style.detail", string="Header Style")
    body_style_excel = fields.Many2one(comodel_name="style.detail", string="Body Style")
    footer_style_excel = fields.Many2one(comodel_name="style.detail", string="Footer Style")

    header_style = fields.Many2one(comodel_name="style.detail", string="Header Style")
    body_style = fields.Many2one(comodel_name="style.detail", string="Body Style")
    footer_style = fields.Many2one(comodel_name="style.detail", string="Footer Style")

    report_id = fields.Many2one(comodel_name="report.report", string="Report")


class ReportDesign(surya.Sarpam):
    _name = "report.design"

    name = fields.Char(string="Name")
    row_1 = fields.Integer(string="Row 1")
    row_2 = fields.Integer(string="Row 2")
    col_1 = fields.Integer(string="Col 1")
    col_2 = fields.Integer(string="Col 2")
    template = fields.Text(string="Template")

    style_id = fields.Many2one(comodel_name="style.detail", string="Style")
    order_seq = fields.Integer(string="Order")

    report_id = fields.Many2one(comodel_name="report.report", string="Report")


class StyleDetail(surya.Sarpam):
    _name = "style.detail"

    style = fields.Text(string="Style")
    type = fields.Selection(selection=[("css", "CSS"), ("excel", "Excel")])











